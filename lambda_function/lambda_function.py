import json
import urllib.parse
import boto3
import os
import psycopg2
from botocore.exceptions import ClientError
from datetime import datetime
import csv
import botocore
from openpyxl import load_workbook
import io

print('Loading function')

s3 = boto3.client('s3', 'eu-north-1', config=botocore.config.Config(s3={'addressing_style':'path'}))

# Retrieve environment variables for database connection
db_host = os.environ['DB_HOST']
db_name = os.environ['DB_NAME']
db_user = os.environ['DB_USER']
db_password = os.environ['DB_PASSWORD']
db_port = os.environ['DB_PORT']

def process_csv(content):
    # Use csv.reader to parse the CSV content with semicolon delimiter
    csv_reader = csv.reader(content.splitlines(), delimiter=';')
    
    # Skip the header row
    next(csv_reader)
    return csv_reader

def process_excel(content):
    # Load the workbook and get the first sheet
    workbook = load_workbook(filename=io.BytesIO(content), data_only=True)
    sheet = workbook.active
    
    # Convert the sheet rows to a list of lists
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(list(row))
    
    # Skip the header row
    return rows[1:]

def lambda_handler(event, context):
    # Get the object from the event and show its content type
    bucket = event['Records'][0]['s3']['bucket']['name']
    key = urllib.parse.unquote_plus(event['Records'][0]['s3']['object']['key'], encoding='utf-8')

    # Check if the file is inside the 'data' prefix
    if not key.startswith('data/'):
        print(f"File {key} does not match the required pattern.")
        return {
            'statusCode': 400,
            'body': json.dumps(f"File {key} does not match the required pattern.")
        }

    try:
        # Retrieve the object from S3
        response = s3.get_object(Bucket=bucket, Key=key)
        content_type = response['ContentType']
        content = response['Body'].read()

        if content_type == 'text/csv' or key.endswith('.csv'):
            rows = process_csv(content.decode('utf-8'))
        elif content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' or key.endswith('.xlsx'):
            rows = process_excel(content)
        else:
            print(f"File {key} is not a supported format.")
            return {
                'statusCode': 400,
                'body': json.dumps(f"File {key} is not a supported format. Only CSV and Excel files are allowed.")
            }

        # Connect to the PostgreSQL database
        conn = psycopg2.connect(
            dbname=db_name,
            user=db_user,
            password=db_password,
            host=db_host,
            port=db_port
        )
        cursor = conn.cursor()

        # Iterate over rows and insert into the gas_price_data table
        for row in rows:

            insert_stmt = """ insert into test_data (first_name, last_name, age, created_at) 
                            VALUES (%s, %s, %s, now());"""
            cursor.execute(insert_stmt, row)

        conn.commit()
        cursor.close()
        conn.close()
        print("Data imported successfully")

        return {
            'statusCode': 200,
            'body': json.dumps('File processed and data imported successfully')
        }

    except ClientError as e:
        print(e)
        print(f'Error getting object {key} from bucket {bucket}. Make sure they exist and your bucket is in the same region as this function.')
        return {
            'statusCode': 500,
            'body': json.dumps(f"Error: {str(e)}")
        }
    except psycopg2.Error as e:
        print(e)
        print(f'Error executing SQL command: {str(e)}')
        return {
            'statusCode': 500,
            'body': json.dumps(f"Error: {str(e)}")
        }
    except Exception as e:
        print(e)
        print(f'Error processing object {key} from bucket {bucket}.')
        return {
            'statusCode': 500,
            'body': json.dumps(f"Error: {str(e)}")
        }
